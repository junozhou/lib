# Python爬虫实战：单线程、多线程和协程性能对比
# https://mp.weixin.qq.com/s?__biz=MzAxMjUyNDQ5OA==&mid=2653570094&idx=1&sn=af293a9b5b1524984e62d4130627d192&chksm=806e6a93b719e38562935011282be651e096986ac6e247d5041bdc48a65f6c489016359acb2a&scene=27
# 很好的文章包含单线程，多线程（线程池方式），协程的对比
# max_workers 默认 (cpu num *5)
# 一、前言
# 今天我要给大家分享的是如何爬取中农网产品报价数据，并分别用普通的单线程、多线程和协程来爬取，从而对比单线程、多线程和协程在网络爬虫中的性能。
# 目标URL：https://www.zhongnongwang.com/quote/product-htm-page-1.html
# 图片
# 爬取产品品名、最新报价、单位、报价数、报价时间等信息，保存到本地Excel。
# 图片
# 二、爬取测试
# 翻页查看 URL 变化规律：
# https://www.zhongnongwang.com/quote/product-htm-page-1.html
# https://www.zhongnongwang.com/quote/product-htm-page-2.html
# https://www.zhongnongwang.com/quote/product-htm-page-3.html
# https://www.zhongnongwang.com/quote/product-htm-page-4.html
# https://www.zhongnongwang.com/quote/product-htm-page-5.html
# https://www.zhongnongwang.com/quote/product-htm-page-6.html
# 检查网页，可以发现网页结构简单，容易解析和提取数据。
# 图片
# 思路：每一条产品报价信息在 class 为 tb 的 table 标签下的 tbody 下的 tr 标签里，获取到所有 tr 标签的内容，然后遍历，从中提取出每一个产品品名、最新报价、单位、报价数、报价时间等信息。
# -*- coding: UTF-8 -*-
"""
@File    ：demo.py
@Author  ：叶庭云
@CSDN    ：https://yetingyun.blog.csdn.net/
"""
import requests
import logging
from fake_useragent import UserAgent
from lxml import etree


# 日志输出的基本配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')
# 随机产生请求头
ua = UserAgent(verify_ssl=False, path='fake_useragent.json')
url = 'https://www.zhongnongwang.com/quote/product-htm-page-1.html'
# 伪装请求头
headers = {
    "Accept-Encoding": "gzip",  # 使用gzip压缩传输数据让访问更快
    "User-Agent": ua.random
}
# 发送请求  获取响应
rep = requests.get(url, headers=headers)
print(rep.status_code)    # 200
# Xpath定位提取数据
html = etree.HTML(rep.text)
items = html.xpath('/html/body/div[10]/table/tr[@align="center"]')
logging.info(f'该页有多少条信息：{len(items)}')  # 一页有20条信息
# 遍历提取出数据
for item in items:
    name = ''.join(item.xpath('.//td[1]/a/text()'))  # 品名
    price = ''.join(item.xpath('.//td[3]/text()'))   # 最新报价
    unit = ''.join(item.xpath('.//td[4]/text()'))    # 单位
    nums = ''.join(item.xpath('.//td[5]/text()'))    # 报价数
    time_ = ''.join(item.xpath('.//td[6]/text()'))   # 报价时间
    logging.info([name, price, unit, nums, time_])
# 运行结果如下：
# 图片
# 可以成功爬取到数据，接下来分别用普通的单线程、多线程和协程来爬取 50 页的数据、保存到Excel。
# 三、单线程爬虫
# -*- coding: UTF-8 -*-
"""
@File    ：单线程.py
@Author  ：叶庭云
@CSDN    ：https://yetingyun.blog.csdn.net/
"""
import requests
import logging
from fake_useragent import UserAgent
from lxml import etree
import openpyxl
from datetime import datetime

# 日志输出的基本配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')
# 随机产生请求头
ua = UserAgent(verify_ssl=False, path='fake_useragent.json')
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(['品名', '最新报价', '单位', '报价数', '报价时间'])
start = datetime.now()

for page in range(1, 51):
    # 构造URL
    url = f'https://www.zhongnongwang.com/quote/product-htm-page-{page}.html'
    # 伪装请求头
    headers = {
        "Accept-Encoding": "gzip",  # 使用gzip压缩传输数据让访问更快
        "User-Agent": ua.random
    }
    # 发送请求  获取响应
    rep = requests.get(url, headers=headers)
    # print(rep.status_code)
    # Xpath定位提取数据
    html = etree.HTML(rep.text)
    items = html.xpath('/html/body/div[10]/table/tr[@align="center"]')
    logging.info(f'该页有多少条信息：{len(items)}')  # 一页有20条信息
    # 遍历提取出数据
    for item in items:
        name = ''.join(item.xpath('.//td[1]/a/text()'))  # 品名
        price = ''.join(item.xpath('.//td[3]/text()'))   # 最新报价
        unit = ''.join(item.xpath('.//td[4]/text()'))    # 单位
        nums = ''.join(item.xpath('.//td[5]/text()'))    # 报价数
        time_ = ''.join(item.xpath('.//td[6]/text()'))   # 报价时间
        sheet.append([name, price, unit, nums, time_])
        logging.info([name, price, unit, nums, time_])


wb.save(filename='data1.xlsx')
delta = (datetime.now() - start).total_seconds()
logging.info(f'用时：{delta}s')
# 运行结果如下：
# 图片
#
# 单线程爬虫必须上一个页面爬取完成才能继续爬取，还可能受当时网络状态影响，用时48.528703s，才将数据爬取完，速度比较慢。
#
# 四、多线程爬虫
# -*- coding: UTF-8 -*-
"""
@File    ：多线程.py
@Author  ：叶庭云
@CSDN    ：https://yetingyun.blog.csdn.net/
"""
import requests
import logging
from fake_useragent import UserAgent
from lxml import etree
import openpyxl
from concurrent.futures import ThreadPoolExecutor, wait, ALL_COMPLETED
from datetime import datetime

# 日志输出的基本配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')
# 随机产生请求头
ua = UserAgent(verify_ssl=False, path='fake_useragent.json')
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(['品名', '最新报价', '单位', '报价数', '报价时间'])
start = datetime.now()


def get_data(page):
    # 构造URL
    url = f'https://www.zhongnongwang.com/quote/product-htm-page-{page}.html'
    # 伪装请求头
    headers = {
        "Accept-Encoding": "gzip",    # 使用gzip压缩传输数据让访问更快
        "User-Agent": ua.random
    }
    # 发送请求  获取响应
    rep = requests.get(url, headers=headers)
    # print(rep.status_code)
    # Xpath定位提取数据
    html = etree.HTML(rep.text)
    items = html.xpath('/html/body/div[10]/table/tr[@align="center"]')
    logging.info(f'该页有多少条信息：{len(items)}')  # 一页有20条信息
    # 遍历提取出数据
    for item in items:
        name = ''.join(item.xpath('.//td[1]/a/text()'))   # 品名
        price = ''.join(item.xpath('.//td[3]/text()'))    # 最新报价
        unit = ''.join(item.xpath('.//td[4]/text()'))     # 单位
        nums = ''.join(item.xpath('.//td[5]/text()'))     # 报价数
        time_ = ''.join(item.xpath('.//td[6]/text()'))    # 报价时间
        sheet.append([name, price, unit, nums, time_])
        logging.info([name, price, unit, nums, time_])


def run():
    # 爬取1-50页
    # max_workers 默认 (cpu num *5)个线程
    with ThreadPoolExecutor(max_workers=6) as executor:
        future_tasks = [executor.submit(get_data, i) for i in range(1, 51)]
        # wait 方法接收3个参数，等待的任务序列、超时时间以及等待条件。等待条件 reture_when 默认为 ALL_COMPLETED，表明要等待所有的任务都结束。可以看到运行结果中，确实是所有任务都完成了，主线程才打印出 main。等待条件还可以设置为 FIRST_COMPLETED，表示第一个任务完成就停止等待。
        # ————————————————
        # 版权声明：本文为CSDN博主「生活不允许普通人内向」的原创文章，遵循CC 4.0 BY-SA版权协议，转载请附上原文出处链接及本声明。
        # 原文链接：https://blog.csdn.net/xiaoyu_wu/article/details/102820384
        wait(future_tasks, return_when=ALL_COMPLETED)

    wb.save(filename='data2.xlsx')
    delta = (datetime.now() - start).total_seconds()
    print(f'用时：{delta}s')


run()
# 运行结果如下：
# 图片
# 多线程爬虫爬取效率提升非常可观，用时 2.648128s，爬取速度很快。
# 五、异步协程爬虫
# -*- coding: UTF-8 -*-
"""
@File    ：demo1.py
@Author  ：叶庭云
@CSDN    ：https://yetingyun.blog.csdn.net/
"""
import aiohttp
import asyncio
import logging
from fake_useragent import UserAgent
from lxml import etree
import openpyxl
from datetime import datetime

# 日志输出的基本配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')
# 随机产生请求头
ua = UserAgent(verify_ssl=False, path='fake_useragent.json')
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(['品名', '最新报价', '单位', '报价数', '报价时间'])
start = datetime.now()


class Spider(object):
    def __init__(self):
        # self.semaphore = asyncio.Semaphore(6)  # 信号量，有时候需要控制协程数，防止爬的过快被反爬
        self.header = {
                "Accept-Encoding": "gzip",    # 使用gzip压缩传输数据让访问更快
                "User-Agent": ua.random
            }

    async def scrape(self, url):
        # async with self.semaphore:  # 设置最大信号量，有时候需要控制协程数，防止爬的过快被反爬
        session = aiohttp.ClientSession(headers=self.header, connector=aiohttp.TCPConnector(ssl=False))
        response = await session.get(url)
        result = await response.text()
        await session.close()
        return result

    async def scrape_index(self, page):
        url = f'https://www.zhongnongwang.com/quote/product-htm-page-{page}.html'
        text = await self.scrape(url)
        await self.parse(text)

    async def parse(self, text):
        # Xpath定位提取数据
        html = etree.HTML(text)
        items = html.xpath('/html/body/div[10]/table/tr[@align="center"]')
        logging.info(f'该页有多少条信息：{len(items)}')  # 一页有20条信息
        # 遍历提取出数据
        for item in items:
            name = ''.join(item.xpath('.//td[1]/a/text()'))  # 品名
            price = ''.join(item.xpath('.//td[3]/text()'))  # 最新报价
            unit = ''.join(item.xpath('.//td[4]/text()'))  # 单位
            nums = ''.join(item.xpath('.//td[5]/text()'))  # 报价数
            time_ = ''.join(item.xpath('.//td[6]/text()'))  # 报价时间
            sheet.append([name, price, unit, nums, time_])
            logging.info([name, price, unit, nums, time_])

    def main(self):
        # 50页的数据
        scrape_index_tasks = [asyncio.ensure_future(self.scrape_index(page)) for page in range(1, 51)]
        loop = asyncio.get_event_loop()
        tasks = asyncio.gather(*scrape_index_tasks)
        loop.run_until_complete(tasks)


if __name__ == '__main__':
    spider = Spider()
    spider.main()
    wb.save('data3.xlsx')
    delta = (datetime.now() - start).total_seconds()
    print("用时：{:.3f}s".format(delta))
# 运行结果如下：
# 图片
# 而到了协程异步爬虫，爬取速度更快，嗖的一下，用时 0.930s 就爬取完 50 页数据，aiohttp + asyncio 异步爬虫竟恐怖如斯。异步爬虫在服务器能承受高并发的前提下增加并发数量，爬取效率提升是非常可观的，比多线程还要快一些。
# 三种爬虫都将 50 页的数据爬取下来保存到了本地，结果如下：
# 图片
# 六、总结回顾
# 今天我演示了简单的单线程爬虫、多线程爬虫和协程异步爬虫。可以看到一般情况下异步爬虫速度最快，多线程爬虫略慢一点，单线程爬虫速度较慢，必须上一个页面爬取完成才能继续爬取。
# 但协程异步爬虫相对来说并不是那么好编写，数据抓取无法使用 request 库，只能使用aiohttp，而且爬取数据量大时，异步爬虫需要设置最大信号量来控制协程数，防止爬的过快被反爬。所以在实际编写 Python 爬虫时，我们一般都会使用多线程爬虫来提速，但必须注意的是网站都有 ip 访问频率限制，爬的过快可能会被封ip，所以一般我们在多线程提速的同时可以使用代理 ip 来并发地爬取数据。
# 多线程(multithreading)：是指从软件或者硬件上实现多个线程并发执行的技术。具有多线程能力的计算机因有硬件支持而能够在同一时间执行多于一个线程，进而提升整体处理性能。具有这种能力的系统包括对称多处理机、多核心处理器以及芯片级多处理或同时多线程处理器。在一个程序中，这些独立运行的程序片段叫作 "线程" (Thread），利用它编程的概念就叫作 "多线程处理"。
# 异步(asynchronous)：为完成某个任务，不同程序单元之间过程中无需通信协调，也能完成任务的方式，不相关的程序单元之间可以是异步的。例如，爬虫下载网页。调度程序调用下载程序后，即可调度其他任务，而无需与该下载任务保持通信以协调行为。不同网页的下载、保存等操作都是无关的，也无需相互通知协调。这些异步操作的完成时刻并不确定。简言之，异步意味着无序。
# 协程(coroutine)，又称微线程、纤程，协程是一种用户态的轻量级线程。协程拥有自己的寄存器上下文和栈。协程调度切换时，将寄存器上下文和栈保存到其他地方，在切回来的时候，恢复先前保存的寄存器上下文和栈。因此协程能保留上一次调用时的状态，即所有局部状态的一个特定组合，每次过程重入时，就相当于进入上一次调用的状态。协程本质上是个单进程，协程相对于多进程来说，无需线程上下文切换的开销，无需原子操作锁定及同步的开销，编程模型也非常简单。我们可以使用协程来实现异步操作，比如在网络爬虫场景下，我们发出一个请求之后，需要等待一定的时间才能得到响应，但其实在这个等待过程中，程序可以干许多其他的事情，等到响应得到之后才切换回来继续处理，这样可以充分利用 CPU 和其他资源，这就是协程的优势。
# 作者：叶庭云
# CSDN：https://yetingyun.blog.csdn.net/
# 热爱可抵岁月漫长，发现求知的乐趣，在不断总结和学习中进步，与诸君共勉。